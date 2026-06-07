import openpyxl
import sys

path = r"C:\Users\leand\Downloads\Libro 1.xlsx"

# Load with formulas (data_only=False keeps formulas)
wb = openpyxl.load_workbook(path, data_only=False)

print("=== SHEETS IN WORKBOOK ===\n")
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"{i+1}. '{name}'  ->  dims={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}")

print("\n\n=== PREVIEW EACH SHEET (first 6 rows, up to 12 cols) ===")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n--- '{name}' (max_row={ws.max_row}, max_col={ws.max_column}) ---")
    for r in range(1, min(ws.max_row, 6) + 1):
        row_vals = []
        for c in range(1, min(ws.max_column, 12) + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                v = ""
            row_vals.append(str(v)[:18])
        print(f"  R{r}: " + " | ".join(row_vals))
