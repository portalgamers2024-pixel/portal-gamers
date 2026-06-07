import openpyxl

path = r"C:\Users\leand\Downloads\Libro 1.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)
ws = wb["CALCULADORA"]

# Find the real extent of meaningful data (non-empty rows)
last_real_row = 0
for r in range(1, ws.max_row + 1):
    has_val = False
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != "":
            has_val = True
            break
    if has_val:
        last_real_row = r

print(f"max_row reported: {ws.max_row}")
print(f"last row with actual content: {last_real_row}")
print(f"max_col: {ws.max_column}\n")

print("=== CALCULADORA full dump (rows 1..%d, cols A..N) ===" % min(last_real_row, 60))
for r in range(1, min(last_real_row, 60) + 1):
    cells = []
    for c in range(1, 15):
        cell = ws.cell(row=r, column=c)
        v = cell.value
        if v is None:
            v = ""
        # mark formulas
        s = str(v)
        cells.append(s[:16])
    # only print rows that have something
    if any(x.strip() for x in cells):
        col_letter = openpyxl.utils.get_column_letter
        print(f"R{r:>3}: " + " | ".join(cells))
