import openpyxl, warnings
warnings.simplefilter("ignore")
from openpyxl.utils import get_column_letter

path = r"C:\Users\leand\Downloads\Libro 1.xlsx"
wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
ws = wb["CALCULADORA"]

print("Scanning rows 42..end for any non-empty cell...")
found = []
r = 0
for row in ws.iter_rows(values_only=False):
    r += 1
    if r < 42:
        continue
    for cell in row:
        v = cell.value
        if v is not None and str(v).strip() != "":
            found.append((r, get_column_letter(cell.column), str(v)[:30]))
print(f"non-empty cells below row 41: {len(found)}")
for r,c,v in found[:60]:
    print(f"  {c}{r} = {v}")
if len(found) > 60:
    print(f"  ... and {len(found)-60} more")
    # show the max row
    print(f"  last populated row: {found[-1][0]}")
wb.close()
