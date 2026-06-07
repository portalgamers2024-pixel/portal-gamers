import openpyxl, warnings
warnings.simplefilter("ignore")
from openpyxl.utils import get_column_letter

path = r"C:\Users\leand\Downloads\Libro 1.xlsx"
wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
ws = wb["CALCULADORA"]

# 1) Find true last meaningful row across ALL 40 cols, and which cols are ever used
last_row = 0
cols_used = set()
rcount = 0
for row in ws.iter_rows(values_only=False):
    rcount += 1
    any_val = False
    for cell in row:
        v = cell.value
        if v is not None and str(v).strip() != "":
            any_val = True
            cols_used.add(cell.column)
    if any_val:
        last_row = cell.row if False else rcount
print(f"rows iterated: {rcount}")
print(f"true last meaningful row: {last_row}")
print(f"columns ever used: {sorted([get_column_letter(c) for c in cols_used], key=lambda x:(len(x),x))}")
wb.close()

# 2) Dump all meaningful cells rows 1..last_row (cap 60) across all used cols
wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
ws = wb["CALCULADORA"]
maxc = max(cols_used) if cols_used else 14
print(f"\n=== CALCULADORA non-empty cells (rows 1..{min(last_row,60)}, cols A..{get_column_letter(maxc)}) ===")
r = 0
for row in ws.iter_rows(min_row=1, max_row=min(last_row,60), max_col=maxc, values_only=False):
    r += 1
    parts = []
    for cell in row:
        v = cell.value
        if v is not None and str(v).strip() != "":
            parts.append(f"{get_column_letter(cell.column)}{r}={str(v)[:22]}")
    if parts:
        print("  " + "  ".join(parts))
wb.close()
