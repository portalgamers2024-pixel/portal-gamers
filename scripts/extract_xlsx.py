"""
Extract all tabs from 'Libro 1.xlsx' into a migration JSON.
- Preserves formulas (strings starting with '=').
- Handles ArrayFormula objects.
- Formats datetimes to strings Sheets will parse.
- Bounds each tab to its real data region (drops empty fill-down tails).
- Rewrites intra-workbook sheet references in formulas per the rename map,
  so references survive the tab renames.
"""
import json, datetime, re, warnings, sys
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter
warnings.simplefilter("ignore")
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

SRC = r"C:\Users\leand\Downloads\Libro 1.xlsx"
OUT = r"C:\Users\leand\portal-gamers\data\migration.json"

# source tab -> target tab name
RENAME = {
    "CALCULADORA": "💰 Precios",
    "DOFUS": "📝 Ventas DOFUS",
    "ALBION": "📝 Ventas ALBION",
    "WOW": "📝 Ventas WOW",
    "STOCKS": "📦 Stock Y Cuentas",
    "Formulado": "🔑 Códigos",
    "Construccion 2": "📋 Registro Códigos",
    "PUBLICACIONES": "📢 Publicaciones",
    "Streaming": "🎬 Streaming",
    "Cuentas Paginas y demas": "🔐 Cuentas Admin",
}

# Hard caps per tab to avoid uploading thousands of empty formula rows.
# We bound by "last row with a literal value", but never exceed these.
MAXROW_CAP = {
    "CALCULADORA": 45,   # real data ends ~41
}
DEFAULT_CAP = 4000


def rewrite_refs(formula: str) -> str:
    """Replace old sheet-name references with quoted new names."""
    out = formula
    for src, tgt in RENAME.items():
        new_ref = f"'{tgt}'!"
        # quoted form: 'Src Name'!  -> 'Target'!
        out = out.replace(f"'{src}'!", new_ref)
        # bare form: Src!  (only when src has no spaces it appears bare)
        if " " not in src:
            out = re.sub(rf"(?<![A-Za-z0-9_]){re.escape(src)}!", new_ref, out)
    return out


def cellval(cell):
    v = cell.value
    if v is None:
        return None
    if isinstance(v, ArrayFormula):
        txt = v.text or ""
        if not txt.startswith("="):
            txt = "=" + txt
        return rewrite_refs(txt)
    if isinstance(v, str):
        if v.startswith("="):
            return rewrite_refs(v)
        return v
    if isinstance(v, datetime.datetime):
        # drop midnight time component for cleanliness
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime.timedelta):
        return str(v)
    # numbers / bools pass through
    return v


def is_formula(x):
    return isinstance(x, str) and x.startswith("=")


def extract_sheet(ws, src_name):
    cap = MAXROW_CAP.get(src_name, DEFAULT_CAP)
    # First pass: load bounded grid
    max_col = ws.max_column
    grid = []
    last_literal_row = 0   # last row containing a non-formula, non-empty value
    last_any_row = 0       # last row containing anything (incl formulas)
    r = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, cap), max_col=max_col, values_only=False):
        r += 1
        out_row = []
        has_literal = False
        has_any = False
        for cell in row:
            cv = cellval(cell)
            out_row.append(cv)
            if cv is not None and str(cv).strip() != "":
                has_any = True
                if not is_formula(cv):
                    has_literal = True
        grid.append(out_row)
        if has_any:
            last_any_row = r
        if has_literal:
            last_literal_row = r
    bound = max(last_literal_row, 1)
    # include a couple trailing formula rows if they immediately follow data
    if last_any_row > bound and last_any_row - bound <= 3:
        bound = last_any_row
    grid = grid[:bound]
    # trim trailing all-empty columns
    width = 0
    for row in grid:
        for ci in range(len(row) - 1, -1, -1):
            if row[ci] is not None and str(row[ci]).strip() != "":
                width = max(width, ci + 1)
                break
    grid = [row[:width] for row in grid]
    return grid, bound, width


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False, read_only=True)
    result = {"rename": RENAME, "tabs": {}}
    print("Extracting tabs:\n")
    for src in RENAME:
        if src not in wb.sheetnames:
            print(f"  !! source tab '{src}' not found, skipping")
            continue
        ws = wb[src]
        grid, rows, cols = extract_sheet(ws, src)
        # count non-empty cells
        nonempty = sum(1 for row in grid for c in row if c is not None and str(c).strip() != "")
        formulas = sum(1 for row in grid for c in row if is_formula(c))
        result["tabs"][src] = {
            "target": RENAME[src],
            "rows": rows,
            "cols": cols,
            "values": grid,
        }
        print(f"  {src:28s} -> {RENAME[src]:22s}  rows={rows:<5} cols={cols:<3} cells={nonempty:<5} formulas={formulas}")
    wb.close()
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False)
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
